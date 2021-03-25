import openpyxl


wb = openpyxl.load_workbook("/Users/skpatro/PycharmProjects/SampleSelPython/test_data.xlsx")
sh = wb['Sheet1']

# fetch single cell value
# print(sh.cell(1, 1).value)

# print(sh['A1'].value)

# fetch multiple rows

"""
for cell in sh['A']:
    print(cell.value)
"""

# print(sh.max_row)  # no. of rows that has data

# print(sh.max_column)  # no. of columns that has data

""""
for i in range(1, sh.max_row+1):
    for cell in sh[i]:
        print(cell.value)
"""

""""
for row in sh.iter_rows(1, sh.max_row):
    for cell in row:
        print(cell.value)
"""

"""
sheet_cells = []  # list of tuples to hold multiple rows data
for row in sh.iter_rows():
    row_cells = []
    for cell in row:
        row_cells.append(cell.value)
    sheet_cells.append(tuple(row_cells))
print(sheet_cells)

for i in range(1, len(sheet_cells)):
    print(sheet_cells[i][0] + " | " + str(sheet_cells[i][1]) + " | " + sheet_cells[i][2])
"""

# wb.create_sheet('data')
sh1 = wb['data']

# sh1.cell(1, 1).value = "qavbox"
# sh1['A1'].value = "QAVBOX"
# print(sh1.cell(1, 1).value)

testdata = [('username', 'password', 'type'), ('abc', 123, 'valid'), ('def', 456, 'invalid')]
for item in testdata:
    sh1.append(item)

testdata1 = ("qavbox", "qavalidation.com")
sh1.append(testdata1)
wb.save("/Users/skpatro/PycharmProjects/SampleSelPython/test_data.xlsx")
