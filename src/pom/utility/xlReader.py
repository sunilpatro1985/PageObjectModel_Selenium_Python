import openpyxl

def load_excel(filepath, sheet):
    global wb
    global sh
    wb = openpyxl.load_workbook(filepath)
    sh = wb[sheet]


def get_cell_data(rowNo, colNo):
    return sh.cell(rowNo, colNo).value


def get_data_as_list_tuples():
    sheet_cells = []
    for i in range(1, sh.max_row):
        row_cells = []
        for cell in sh[i+1]:
            row_cells.append(cell.value)
        sheet_cells.append(tuple(row_cells))
    return sheet_cells

"""
def get_max_rows():


def get_max_columns():
"""
